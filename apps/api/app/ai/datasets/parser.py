"""Spreadsheet parser (Capability 1, Slice 1b) — PURE, in-process.

`parse_spreadsheet(bytes, kind=...)` turns an .xlsx / .csv into a typed schema
(`DatasetColumnSpec`s) + coerced JSONB-ready rows. No DB, no I/O beyond reading
the given bytes.

Detection:
  - header row (`header_row_index`, default 0) → per header: `name` = original,
    `key` = normalized slug (lowercase, non-alphanumeric → `_`, collisions get a
    `_2`/`_3` suffix), `position` = ordinal.
  - data_type per column, by sampling its NON-EMPTY values: all parse as
    int/float → number; else all parse as a date → date; else all in
    {true/false/yes/no/y/n} → boolean; else text.
  - semantic_role by header name (case-insensitive substring): status/state/stage
    → status; owner/assignee/assigned/responsible/rep → owner; priority/severity
    → priority; due/deadline/target date/close date → due_date; else none.
  - each data row → dict keyed by column.key, values coerced to the column type
    (number → int/float, date → ISO-8601 string, boolean → bool, else trimmed
    string; empty → None). Fully-empty rows are skipped.

Guardrails: row and column caps (raise `SpreadsheetTooLargeError`), plus clear
errors on an empty sheet / missing header. See `app/ai/datasets/ingest.py` for
the note on parse ISOLATION (in-process here; subprocess hardening is a follow-up).

SECURITY: `openpyxl` is opened in `read_only` mode; the caps bound how much of an
untrusted file is materialised.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Generator, Sequence
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime

from app.application.ports.repositories import DatasetColumnSpec
from app.shared.exceptions.datasets import (
    EmptySpreadsheetError,
    NoHeaderError,
    SpreadsheetDecodeError,
    SpreadsheetTooLargeError,
)

MAX_ROWS = 50_000
MAX_COLUMNS = 200

_TRUE_TOKENS = frozenset({"true", "yes", "y"})
_FALSE_TOKENS = frozenset({"false", "no", "n"})

# semantic_role keyword map, checked in this order (first match wins).
_ROLE_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("status", ("status", "state", "stage")),
    ("owner", ("owner", "assignee", "assigned", "responsible", "rep")),
    ("priority", ("priority", "severity")),
    ("due_date", ("due", "deadline", "target date", "close date")),
)

# Extra date formats tried after ISO-8601.
_DATE_FORMATS = ("%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y", "%d %b %Y")

# A raw cell value as read from the source (native for xlsx, str for csv).
Cell = object


@dataclass(frozen=True, slots=True)
class ParsedSpreadsheet:
    columns: list[DatasetColumnSpec]
    rows: list[dict[str, object]]


# --- value predicates / coercions --------------------------------------------


def _is_empty(v: Cell) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _as_number(v: Cell) -> float | None:
    if isinstance(v, bool):  # bool is an int subclass — never a number column
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def _as_date(v: Cell) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        try:
            return date.fromisoformat(s)
        except ValueError:
            pass
        try:
            return datetime.fromisoformat(s).date()
        except ValueError:
            pass
        for fmt in _DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _as_bool(v: Cell) -> bool | None:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        s = v.strip().lower()
        if s in _TRUE_TOKENS:
            return True
        if s in _FALSE_TOKENS:
            return False
    return None


def _infer_type(values: Sequence[Cell]) -> str:
    """Infer a column's data_type from its NON-EMPTY sample values."""
    if not values:
        return "text"
    if all(_as_number(v) is not None for v in values):
        return "number"
    if all(_as_date(v) is not None for v in values):
        return "date"
    if all(_as_bool(v) is not None for v in values):
        return "boolean"
    return "text"


def _coerce(v: Cell, data_type: str) -> object:
    if _is_empty(v):
        return None
    if data_type == "number":
        n = _as_number(v)
        if n is None:
            return str(v).strip()
        return int(n) if n.is_integer() else n
    if data_type == "date":
        d = _as_date(v)
        return d.isoformat() if d is not None else str(v).strip()
    if data_type == "boolean":
        b = _as_bool(v)
        return b if b is not None else str(v).strip()
    return str(v).strip()


# --- header handling ---------------------------------------------------------


def _slugify(header: str, position: int) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", header.strip().lower()).strip("_")
    return s or f"column_{position + 1}"


def _semantic_role(header: str) -> str:
    h = header.lower()
    for role, keywords in _ROLE_KEYWORDS:
        if any(k in h for k in keywords):
            return role
    return "none"


def _header_text(cell: Cell) -> str:
    return "" if cell is None else str(cell).strip()


def _dedupe_keys(headers: list[str]) -> list[str]:
    """Normalize each header to a slug; collisions get `_2`, `_3`, … suffixes."""
    counts: dict[str, int] = {}
    keys: list[str] = []
    for position, header in enumerate(headers):
        base = _slugify(header, position)
        n = counts.get(base, 0) + 1
        counts[base] = n
        key = base if n == 1 else f"{base}_{n}"
        while key in counts and key != base:  # guard a literal clash with a slug
            n += 1
            counts[base] = n
            key = f"{base}_{n}"
        counts.setdefault(key, 1)
        keys.append(key)
    return keys


# --- row sources -------------------------------------------------------------


def _xlsx_rows(data: bytes, sheet_name: str | None) -> Generator[list[Cell], None, None]:
    # Lazy import: only an xlsx ingest pays for openpyxl.
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise EmptySpreadsheetError(f"sheet {sheet_name!r} not found")
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0] if wb.worksheets else None
            if ws is None:
                raise EmptySpreadsheetError("workbook has no sheets")
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        wb.close()


def _csv_rows(data: bytes) -> Generator[list[Cell], None, None]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as e:
        raise SpreadsheetDecodeError("could not decode CSV as UTF-8") from e
    for row in csv.reader(io.StringIO(text)):
        yield list(row)


# --- entry point -------------------------------------------------------------


def parse_spreadsheet(
    data: bytes,
    *,
    kind: str,
    sheet_name: str | None = None,
    header_row_index: int = 0,
    max_rows: int = MAX_ROWS,
    max_columns: int = MAX_COLUMNS,
) -> ParsedSpreadsheet:
    """Parse `data` (kind='xlsx'|'csv') into columns + coerced rows.

    Raises SpreadsheetTooLargeError / EmptySpreadsheetError / NoHeaderError /
    SpreadsheetDecodeError. `kind` is trusted (the route resolves + validates it).
    """
    if kind == "xlsx":
        source = _xlsx_rows(data, sheet_name)
    elif kind == "csv":
        source = _csv_rows(data)
    else:  # pragma: no cover - the route gates kind before calling
        raise ValueError(f"unsupported spreadsheet kind: {kind!r}")

    if header_row_index < 0:
        raise NoHeaderError("header_row_index must be >= 0")

    with closing(source):
        indexed = enumerate(source)

        # Advance to (and capture) the header row.
        header_row: list[Cell] | None = None
        for i, row in indexed:
            if i < header_row_index:
                continue
            header_row = row
            break
        if header_row is None:
            raise EmptySpreadsheetError("spreadsheet has no header row")

        headers = [_header_text(c) for c in header_row]
        # Trim trailing empty header cells (openpyxl pads short rows).
        while headers and headers[-1] == "":
            headers.pop()
        if not headers or all(h == "" for h in headers):
            raise NoHeaderError("header row is empty")
        if len(headers) > max_columns:
            raise SpreadsheetTooLargeError(
                f"{len(headers)} columns exceeds the {max_columns}-column limit"
            )

        n_cols = len(headers)
        keys = _dedupe_keys(headers)

        # Collect non-empty data rows (truncated/padded to the header width),
        # capping the count before materialising an unbounded sheet.
        data_rows: list[list[Cell]] = []
        for _i, row in indexed:
            cells: list[Cell] = [row[c] if c < len(row) else None for c in range(n_cols)]
            if all(_is_empty(c) for c in cells):
                continue
            data_rows.append(cells)
            if len(data_rows) > max_rows:
                raise SpreadsheetTooLargeError(
                    f"more than {max_rows} data rows exceeds the row limit"
                )

    # Infer each column's type from its non-empty values, then build the schema.
    columns: list[DatasetColumnSpec] = []
    types: list[str] = []
    for pos in range(n_cols):
        col_values = [r[pos] for r in data_rows]
        non_empty = [v for v in col_values if not _is_empty(v)]
        data_type = _infer_type(non_empty)
        types.append(data_type)
        columns.append(
            DatasetColumnSpec(
                name=headers[pos] or keys[pos],
                key=keys[pos],
                data_type=data_type,
                position=pos,
                nullable=(len(non_empty) < len(col_values)) if data_rows else True,
                semantic_role=_semantic_role(headers[pos]),
            )
        )

    rows: list[dict[str, object]] = [
        {keys[pos]: _coerce(raw[pos], types[pos]) for pos in range(n_cols)} for raw in data_rows
    ]
    return ParsedSpreadsheet(columns=columns, rows=rows)
